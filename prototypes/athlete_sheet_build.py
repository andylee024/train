#!/usr/bin/env python3
"""
Prototype: athlete-facing sheet rendered as .xlsx.

Spec:    docs/product/athlete-sheet-format.md
Example: Andy Lee, Block 1 Wk 2 — today = Fri May 15, 2026
Output:  prototypes/athlete_sheet_andy_wk2.xlsx

Run: python3 prototypes/athlete_sheet_build.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "athlete_sheet_andy_wk2.xlsx"

# ── Palette ────────────────────────────────────────────────────────────────
NAVY = "1A1A2E"
SECTION = "F0F4FF"
TODAY = "FEF3C7"
DONE = "E6F4EA"
WARN = "FFF3E0"
B1_TINT = "DBEAFE"
HEAD_LIGHT = "F1F5F9"
PR_GREEN = "DCFCE7"
GREY_TEXT = "666666"

# ── Fonts ──────────────────────────────────────────────────────────────────
T = "Arial"
TITLE = Font(name=T, size=16, bold=True, color="FFFFFF")
SUB = Font(name=T, size=11, color="FFFFFF")
H1 = Font(name=T, size=13, bold=True)
H2 = Font(name=T, size=11, bold=True)
BODY = Font(name=T, size=10)
BODY_BOLD = Font(name=T, size=10, bold=True)
BODY_DIM = Font(name=T, size=10, color=GREY_TEXT)
MUTED = Font(name=T, size=9, color=GREY_TEXT, italic=True)
TAG_DONE = Font(name=T, size=10, bold=True, color="166534")
TAG_TODAY = Font(name=T, size=10, bold=True, color="92400E")
TAG_PENDING = Font(name=T, size=10, color=GREY_TEXT)

# ── Fills ──────────────────────────────────────────────────────────────────
F_NAVY = PatternFill("solid", fgColor=NAVY)
F_SECTION = PatternFill("solid", fgColor=SECTION)
F_TODAY = PatternFill("solid", fgColor=TODAY)
F_DONE = PatternFill("solid", fgColor=DONE)
F_WARN = PatternFill("solid", fgColor=WARN)
F_B1 = PatternFill("solid", fgColor=B1_TINT)
F_HEAD = PatternFill("solid", fgColor=HEAD_LIGHT)
F_PR = PatternFill("solid", fgColor=PR_GREEN)

# ── Borders / Alignment ────────────────────────────────────────────────────
_thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

L = Alignment(horizontal="left", vertical="center", wrap_text=True)
LT = Alignment(horizontal="left", vertical="top", wrap_text=True)
C = Alignment(horizontal="center", vertical="center")
R = Alignment(horizontal="right", vertical="center")


# ── Helpers ────────────────────────────────────────────────────────────────
def widths(ws, *vals):
    for i, w in enumerate(vals, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def merge_set(ws, row, col, end_col, value, font=BODY, fill=None, align=L, height=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align
    if end_col > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
        # Apply fill to every cell in the merged range so banding looks right.
        if fill:
            for c in range(col + 1, end_col + 1):
                ws.cell(row=row, column=c).fill = fill
    if height:
        ws.row_dimensions[row].height = height
    return cell


def title_bar(ws, row, left_text, right_text, span):
    """Two-line title: navy bar with title + subtitle."""
    merge_set(ws, row, 1, span, left_text, font=TITLE, fill=F_NAVY, align=L, height=30)
    merge_set(ws, row + 1, 1, span, right_text, font=SUB, fill=F_NAVY, align=L, height=22)


def section(ws, row, text, span):
    merge_set(ws, row, 1, span, text, font=H1, fill=F_SECTION, align=L, height=22)


def kv_row(ws, row, label, value, label_col=1, value_col=2, value_span=None):
    ws.cell(row=row, column=label_col, value=label).font = BODY_BOLD
    vc = ws.cell(row=row, column=value_col, value=value)
    vc.font = BODY
    vc.alignment = L
    if value_span and value_span > value_col:
        ws.merge_cells(start_row=row, start_column=value_col, end_row=row, end_column=value_span)


def table_header(ws, row, headers, fill=F_HEAD, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = BODY_BOLD
        c.fill = fill
        c.alignment = C
        c.border = BORDER


def table_row(ws, row, values, start_col=1, fill=None, bold_first=False, align=None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = BODY_BOLD if (bold_first and i == 0) else BODY
        c.alignment = align or (L if isinstance(v, str) else C)
        c.border = BORDER
        if fill:
            c.fill = fill


# ══════════════════════════════════════════════════════════════════════════
#                              TAB 1 — HOME
# ══════════════════════════════════════════════════════════════════════════
def build_home(ws):
    widths(ws, 4, 22, 16, 14, 14, 14, 12)
    span = 7

    # Title bar
    title_bar(
        ws, 1,
        "ANDY LEE  —  SUMMER 2026 DUNK ARC                        Wk 2 of 18",
        "Block 1: Power Conversion + Upper Build                  113 days to test day",
        span,
    )

    # ── TODAY ──
    row = 4
    section(ws, row, "▸  TODAY   (Fri May 15)", span); row += 1
    merge_set(ws, row, 1, span, "  Upper Vertical — Pull-Up + Dips  ·  ~60 min",
              font=H2, fill=F_TODAY, align=L, height=22); row += 1
    merge_set(ws, row, 1, span,
              "  Pull-up +15  •  Dips +15  •  Chest-supp Row  •  Landmine Press  •  DB Curl  •  Tri Pushdown",
              font=BODY, fill=F_TODAY, align=L, height=20); row += 1
    merge_set(ws, row, 1, span,
              "  → Open the “This Week” tab to log,   or text your sets to your coach",
              font=MUTED, fill=F_TODAY, align=L, height=20); row += 2

    # ── THIS WEEK ──
    section(ws, row, "▸  THIS WEEK", span); row += 1
    week = [
        ("Sun  May 10", "Lower Strength (LIGHT) + Posterior",        "✓ Done",     "done"),
        ("Mon  May 11", "Upper Horizontal — Bench + Row",            "✓ Done",     "done"),
        ("Tue  May 12", "Speed-Strength B — Olympic + Sprints",      "✓ Done",     "done"),
        ("Wed  May 13", "Recovery + Deep Flex + Pull-Up Volume",     "✓ Done",     "done"),
        ("Thu  May 14", "Speed-Strength A — Power Clean + Squat",    "✓ Done",     "done"),
        ("Fri  May 15", "Upper Vertical — Pull-Up + Dips",           "○ Today",    "today"),
        ("Sat  May 16", "JUMP LAB — Sprint + Approach + Depth",      "○ Tomorrow", "pending"),
    ]
    for day, sess, status, kind in week:
        ws.cell(row=row, column=1, value="").fill = (
            F_DONE if kind == "done" else F_TODAY if kind == "today" else F_HEAD
        )
        d = ws.cell(row=row, column=2, value=day);  d.font = BODY_BOLD; d.alignment = L
        s = ws.cell(row=row, column=3, value=sess); s.font = BODY; s.alignment = L
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=span - 1)
        st = ws.cell(row=row, column=span, value=status)
        st.font = TAG_DONE if kind == "done" else TAG_TODAY if kind == "today" else TAG_PENDING
        st.alignment = R
        # Banding
        fill = F_DONE if kind == "done" else F_TODAY if kind == "today" else None
        if fill:
            for c in range(2, span + 1):
                ws.cell(row=row, column=c).fill = fill
        ws.row_dimensions[row].height = 20
        row += 1
    row += 1

    # ── RECENT PRs ──
    section(ws, row, "▸  RECENT PRs", span); row += 1
    prs = [
        ("Power Clean",      "140 × 3",    "Thu  May 14",  "arc PR"),
        ("Bench Press",      "170 × 6",    "Mon  May 11",  "build wk top set"),
        ("Weighted Pull-Up", "+10 × 5",    "Fri  May 08",  "Wk 1 top set"),
    ]
    table_header(ws, row, ["Lift", "Load × Reps", "Date", "Note"], fill=F_HEAD, start_col=2)
    row += 1
    for lift, load, date, note in prs:
        table_row(ws, row, [lift, load, date, note], start_col=2, fill=F_PR)
        row += 1
    row += 1

    # ── BODYWEIGHT ──
    section(ws, row, "▸  BODYWEIGHT  (7-day rolling avg)", span); row += 1
    bw_cell = ws.cell(row=row, column=2, value="192.1 lb"); bw_cell.font = Font(name=T, size=18, bold=True); bw_cell.alignment = L
    trend = ws.cell(row=row, column=4, value="↓ 0.4 vs last wk"); trend.font = Font(name=T, size=12, bold=True, color="166534"); trend.alignment = L
    target = ws.cell(row=row, column=6, value="target Wk 6: 188 lb"); target.font = MUTED; target.alignment = R
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 26
    row += 2

    # ── COACH NOTE ──
    section(ws, row, "▸  COACH NOTE   (Sun May 10)", span); row += 1
    note = (
        "Wk 1 in the books — clean execution all 6 sessions.  CMJ baseline at 28.5\", "
        "side split −22 cm baseline.  Block 1 emphasis is closing the reactivity gap, "
        "so push intent on speed work and don't grind the squat — it's maintenance.  "
        "Bw down 0.4 on the cut, exactly on the curve."
    )
    c = ws.cell(row=row, column=1, value="  " + note)
    c.font = BODY; c.alignment = LT
    c.fill = F_HEAD
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=span)
    for r in range(row, row + 4):
        for col in range(1, span + 1):
            ws.cell(row=r, column=col).fill = F_HEAD
    ws.row_dimensions[row].height = 22


# ══════════════════════════════════════════════════════════════════════════
#                            TAB 2 — THIS WEEK
# ══════════════════════════════════════════════════════════════════════════
def build_this_week(ws):
    widths(ws, 4, 28, 22, 9, 9, 9, 9, 22)
    span = 8

    title_bar(
        ws, 1,
        "THIS WEEK   —   Wk 2 of 18",
        "Sun May 10  →  Sat May 16, 2026                  Bidirectional: log here OR text your coach",
        span,
    )
    row = 4

    def day_header(r, text, status, kind):
        fill = F_DONE if kind == "done" else F_TODAY if kind == "today" else F_HEAD
        merge_set(ws, r, 1, span - 1, "  " + text, font=H2, fill=fill, align=L, height=24)
        st = ws.cell(row=r, column=span, value=status)
        st.font = TAG_DONE if kind == "done" else TAG_TODAY if kind == "today" else TAG_PENDING
        st.alignment = R
        st.fill = fill

    def exercise_header(r):
        table_header(ws, r, ["#", "Exercise", "Prescribed", "S1", "S2", "S3", "S4", "Notes"])

    def exercise_row(r, num, name, prescribed, sets=("", "", "", ""), notes=""):
        table_row(ws, r, [num, name, prescribed, *sets, notes])

    def footer(r, rpe="___", bw="___", energy="___", notes="___"):
        ws.cell(row=r, column=1, value="").fill = F_HEAD
        merge_set(ws, r, 2, 3,
                  f"RPE overall: {rpe}     Bw post-WO: {bw}     Energy 1-10: {energy}",
                  font=BODY_BOLD, fill=F_HEAD, align=L, height=20)
        merge_set(ws, r, 4, span,
                  f"Session notes:  {notes}",
                  font=BODY, fill=F_HEAD, align=L, height=20)

    # SUN — completed with logged values
    day_header(row, "SUN  May 10  —  Lower Strength (LIGHT) + Posterior", "✓ Done  9:22am", "done"); row += 1
    exercise_header(row); row += 1
    exercise_row(row, 1, "Back Squat (light)",     "1×5 + 2×5 @ 75% (280)",  ("280×5", "265×5", "265×5", ""),      "smooth"); row += 1
    exercise_row(row, 2, "Stiff-Legged Deadlift",  "4×5,5,4,3 @ 195",        ("195×5", "195×5", "195×4", "195×3"), "heavy s4"); row += 1
    exercise_row(row, 3, "Bulgarian Split Squat",  "3×8/leg @ 35 DB",        ("35×8",  "35×8",  "35×8",  ""),      ""); row += 1
    exercise_row(row, 4, "Glute-Ham Raise",        "3×8 @ BW",               ("BW×8",  "BW×8",  "BW×8",  ""),      ""); row += 1
    exercise_row(row, 5, "Pigeon Stretch",         "2×60s/side",             ("60s",   "60s",   "",      ""),      ""); row += 1
    footer(row, rpe="7", bw="192.4", energy="7", notes="SLDL felt heavy on s4 — bump 2.5 next week"); row += 2

    # MON / TUE / WED / THU — collapsed (completed)
    for day, sess in [
        ("MON  May 11  —  Upper Horizontal — Bench + Row",            "✓ Done"),
        ("TUE  May 12  —  Speed-Strength B — Olympic + Sprints",      "✓ Done"),
        ("WED  May 13  —  Recovery + Deep Flex + Pull-Up Volume",     "✓ Done"),
        ("THU  May 14  —  Speed-Strength A — Power Clean + Squat",    "✓ Done"),
    ]:
        day_header(row, day, sess, "done")
        row += 1
        merge_set(ws, row, 2, span, "  (collapsed — tap day header to expand)",
                  font=MUTED, fill=F_DONE, align=L, height=20)
        row += 2

    # FRI — TODAY
    day_header(row, "FRI  May 15  —  Upper Vertical — Pull-Up + Dips", "○ Today", "today"); row += 1
    exercise_header(row); row += 1
    fri = [
        (1, "Weighted Pull-Up",       "4×5 @ +15"),
        (2, "Weighted Dips",          "3×8 @ +15"),
        (3, "Chest-Supported Row",    "4×6-8 @ Mod"),
        (4, "Landmine Press",         "3×10 @ Mod"),
        (5, "Incline DB Curl",        "3×8-10 @ Mod"),
        (6, "Tricep Pushdown",        "3×8-10 @ Mod"),
    ]
    for num, name, presc in fri:
        exercise_row(row, num, name, presc); row += 1
    footer(row, rpe="___", bw="___", energy="___", notes="___ (Shoulder pain 1-10): ___")
    row += 2

    # SAT — tomorrow preview
    day_header(row, "SAT  May 16  —  JUMP LAB — Sprint + Approach + Depth + Loaded Power",
               "○ Tomorrow", "pending"); row += 1
    exercise_header(row); row += 1
    sat = [
        (1, "Sprints",            "3×20yd, 2×30yd @ 95%"),
        (2, "Approach Jumps",     "5×2 @ max intent"),
        (3, "Low-Box Depth Jump", "2×5 @ 18 in"),
        (4, "Barbell Jump",       "4×4 @ 30% (110)"),
        (5, "Hang Power Clean",   "4×3 @ 68% (135)"),
    ]
    for num, name, presc in sat:
        exercise_row(row, num, name, presc); row += 1
    # Sat-specific footer (jump-day metrics)
    ws.cell(row=row, column=1, value="").fill = F_HEAD
    merge_set(ws, row, 2, span,
              "Best jump height: ___    Dunk attempts: ___    Makes: ___    Notes: ___",
              font=BODY_BOLD, fill=F_HEAD, align=L, height=22)


# ══════════════════════════════════════════════════════════════════════════
#                              TAB 3 — PLAN
# ══════════════════════════════════════════════════════════════════════════
def build_plan(ws):
    widths(ws, 4, 22, 14, 16, 16, 16, 14)
    span = 7

    title_bar(
        ws, 1,
        "ARC :  SUMMER 2026 DUNK",
        "May 3, 2026  →  Sep 5, 2026         (18 weeks · 3 blocks · v6)",
        span,
    )
    row = 4

    section(ws, row, "PURPOSE", span); row += 1
    merge_set(ws, row, 1, span,
              "  Close the explosive strength deficit so a strength-dominant athlete can finally express his "
              "force as a dunk.  Build real upper body strength on the side.  Drive the side split as a daily "
              "protocol.  VJ drives the jump methodology; Dylan Shannon drives the upper days and the 4-pillar "
              "lower distribution.",
              font=BODY, fill=F_HEAD, align=LT, height=44)
    row += 2

    section(ws, row, "GOALS", span); row += 1
    table_header(ws, row, ["", "Goal", "Test method", "Deadline"], start_col=1); row += 1
    goals = [
        ("G1", "Dunk a women's basketball on a regulation 10' rim",        "Filmed clean dunk",      "Wk 18 / Sep 5"),
        ("G2", "Bench 1RM 260 lb  AND  weighted pull-up +80 lb × 8 clean", "1RM bench + max reps",   "Wk 18 / Sep 1"),
        ("G3", "Side split distance reduced ≥ 8 cm from baseline",         "Tape measure (weekly)",  "Wk 18 / Sep 1"),
    ]
    for g, desc, test, dl in goals:
        ws.cell(row=row, column=1, value=g).font = BODY_BOLD
        ws.cell(row=row, column=1).alignment = C
        ws.cell(row=row, column=1).fill = F_PR
        ws.cell(row=row, column=1).border = BORDER
        merge_set(ws, row, 2, 4, desc, font=BODY, align=L)
        ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=5, value=test).font = BODY
        ws.cell(row=row, column=5).border = BORDER
        ws.cell(row=row, column=5).alignment = L
        merge_set(ws, row, 6, span, dl, font=BODY, align=L)
        ws.cell(row=row, column=6).border = BORDER
        row += 1
    row += 1

    section(ws, row, "BLOCKS", span); row += 1
    table_header(ws, row, ["#", "Name", "Weeks", "Bet"], start_col=1); row += 1
    # Block 1 — current
    ws.cell(row=row, column=1, value="B1").font = BODY_BOLD
    ws.cell(row=row, column=1).fill = F_B1; ws.cell(row=row, column=1).alignment = C; ws.cell(row=row, column=1).border = BORDER
    merge_set(ws, row, 2, 2, "Power Conversion + Upper Build  ◀  YOU ARE HERE (Wk 2 of 6)",
              font=BODY_BOLD, fill=F_B1, align=L); ws.cell(row=row, column=2).border = BORDER
    c = ws.cell(row=row, column=3, value="W1 – W6"); c.font = BODY; c.fill = F_B1; c.alignment = C; c.border = BORDER
    merge_set(ws, row, 4, span, "Close ESD: jump squats + Olympic + reactive intro; squat on maintenance",
              font=BODY, fill=F_B1, align=L)
    for cc in range(4, span + 1): ws.cell(row=row, column=cc).border = BORDER
    row += 1
    # B2, B3
    other_blocks = [
        ("B2", "Reactive + Dunk Window",       "W7 – W12",  "Depth jumps enter; dunk attempts begin Wk 9; bench peaks ~255"),
        ("B3", "Peak + Realize + Test",        "W13 – W18", "Volume taper; PAP complexes; test bench 260 + pull-up + dunk"),
    ]
    for num, name, weeks, bet in other_blocks:
        ws.cell(row=row, column=1, value=num).font = BODY_BOLD; ws.cell(row=row, column=1).alignment = C; ws.cell(row=row, column=1).border = BORDER
        merge_set(ws, row, 2, 2, name, font=BODY, align=L); ws.cell(row=row, column=2).border = BORDER
        c = ws.cell(row=row, column=3, value=weeks); c.font = BODY; c.alignment = C; c.border = BORDER
        merge_set(ws, row, 4, span, bet, font=BODY, align=L)
        for cc in range(4, span + 1): ws.cell(row=row, column=cc).border = BORDER
        row += 1
    row += 1

    section(ws, row, "TESTS", span); row += 1
    table_header(ws, row, ["Metric", "Baseline", "Wk 6", "Wk 12", "Wk 18", "Method"], start_col=1); row += 1
    tests = [
        ("Standing CMJ",       "28.5\"",      "+1 in",          "+2.5 in",      "+4 in",                  "Best of 3"),
        ("Bounce Depth Jump",  "TBD",         "≥ Standing",     "> Standing",   "Reactive deficit closed", "Best of 3"),
        ("Approach Touch",     "Grab rim",    "Consistent",     "+3 in",        "DUNK",                   "Film + wall mark"),
        ("Back Squat",         "370 lb",      "≥ 360 (held)",   "≥ 360",        "≥ 360",                  "Top set 3RM"),
        ("Bench Press",        "230 lb",      "245 (3RM)",      "255 (1RM)",    "260 (1RM)",              "1RM"),
        ("Weighted Pull-Up",   "+25 × 5",     "+25 × 5 clean",  "+50 × 5",      "+80 × 8",                "Max reps @ target"),
        ("Side Split",         "−22 cm",      "−3 cm",          "−6 cm",        "−8 cm",                  "Tape, weekly"),
        ("Bodyweight",         "192 lb",      "188",            "185",          "183-186 held",           "Daily 7-day avg"),
    ]
    for metric, baseline, w6, w12, w18, method in tests:
        table_row(ws, row, [metric, baseline, w6, w12, w18, method], bold_first=True, align=L)
        row += 1
    row += 1

    section(ws, row, "CONSTRAINTS", span); row += 1
    constraints = [
        "⚠   R shoulder: no BB OHP, no behind-neck press, no jerk catches.  Landmine / DB neutral only.",
        "⚠   L wrist (De Quervain's): no front rack catches.  Use HANG variants on all Olympic lifts; hook grip + straps on pulls.",
        "⚠   Heavy squat must precede jump day by ≥48h (VJ §13).  Sat = jump day → Sun stays LIGHT (top 75–80% only, never push).",
        "⚠   Loaded splits before jump day = NO.  Deep split work Wed ONLY.",
        "⚠   Block 3 = no cut.  Maintenance kcal during peak (VJ §9).",
        "⚠   Patellar pain ≥ 3/10 → drop depth jumps that week.  ≥ 5 → 2-wk plyo deload.",
        "⚠   Squat top set −5% in a week → halve cal deficit.   −8% → pause cut.",
    ]
    for line in constraints:
        merge_set(ws, row, 1, span, "  " + line, font=BODY, fill=F_WARN, align=L, height=22)
        row += 1
    row += 1

    section(ws, row, "PROFILE", span); row += 1
    profile_lines = [
        "6 days/wk training (Sun–Tue, Thu–Sat) · Wed optional flex/recovery · 60–75 min/session, Sat 45 min",
        "32yo · 192 lb · LA · Full commercial gym + sprint surface · Max 6 exercises/session (hard cap)",
        "Strength-dominant, reactivity-deficient (VJ §3 dx) · Squat 1.93× BW (370 lb) · Saturated on strength blocks",
        "Active injuries: R shoulder, L wrist (De Quervain's) — see Constraints above",
        "History: ~1.5 yr toward dunk · Prior fall-off pattern = 2-3 wks gone when programs get complex",
        "Style guides: Vertical Jump Bible + THP (primary) · Dylan Shannon POWERJACKED (secondary)",
    ]
    for line in profile_lines:
        merge_set(ws, row, 1, span, "  " + line, font=BODY, fill=F_HEAD, align=L, height=20)
        row += 1


# ══════════════════════════════════════════════════════════════════════════
#                              TAB 4 — LOG
# ══════════════════════════════════════════════════════════════════════════
def build_log(ws):
    widths(ws, 12, 22, 26, 6, 7, 9, 6, 28)
    span = 8

    title_bar(
        ws, 1,
        "LOG   —   every set, all-time",
        "Source: Supabase exercise_sets   ·   Edits via SMS:  \"actually s3 of bench was 175 not 170\"",
        span,
    )
    row = 4

    table_header(ws, row, ["Date", "Session", "Exercise", "Set", "Reps", "Load", "RPE", "Notes"])
    row += 1

    logs = [
        # Thu 5/14 — Speed-Strength A
        ("Thu 5/14", "Speed-Strength A", "Power Clean",         1, 3, "140",  7, ""),
        ("Thu 5/14", "Speed-Strength A", "Power Clean",         2, 3, "140",  7, ""),
        ("Thu 5/14", "Speed-Strength A", "Power Clean",         3, 3, "140",  7, ""),
        ("Thu 5/14", "Speed-Strength A", "Power Clean",         4, 3, "140",  8, ""),
        ("Thu 5/14", "Speed-Strength A", "Power Clean",         5, 3, "140",  8, "grindy s5"),
        ("Thu 5/14", "Speed-Strength A", "Speed Squat",         1, 2, "215",  7, "fast"),
        ("Thu 5/14", "Speed-Strength A", "Speed Squat",         2, 2, "215",  7, ""),
        ("Thu 5/14", "Speed-Strength A", "Speed Squat",         3, 2, "215",  7, ""),
        ("Thu 5/14", "Speed-Strength A", "Speed Squat",         4, 2, "215",  8, ""),
        ("Thu 5/14", "Speed-Strength A", "Speed Squat",         5, 2, "215",  8, ""),
        ("Thu 5/14", "Speed-Strength A", "Jump Squat",          1, 4, "90",   "-", "intent high"),
        ("Thu 5/14", "Speed-Strength A", "Jump Squat",          2, 4, "90",   "-", ""),
        ("Thu 5/14", "Speed-Strength A", "Jump Squat",          3, 4, "90",   "-", ""),
        ("Thu 5/14", "Speed-Strength A", "Jump Squat",          4, 4, "90",   "-", ""),
        ("Thu 5/14", "Speed-Strength A", "RDL",                 1, 5, "195",  7, ""),
        ("Thu 5/14", "Speed-Strength A", "RDL",                 2, 5, "195",  7, ""),
        ("Thu 5/14", "Speed-Strength A", "RDL",                 3, 4, "195",  8, ""),
        ("Thu 5/14", "Speed-Strength A", "RDL",                 4, 3, "195",  8, ""),
        ("Thu 5/14", "Speed-Strength A", "Pendulum Squat",      1, 8, "mod",  7, ""),
        # Wed 5/13
        ("Wed 5/13", "Recovery + Flex",  "Weighted Pull-Up",    1, 8, "+5",   6, "volume"),
        ("Wed 5/13", "Recovery + Flex",  "Weighted Pull-Up",    2, 8, "+5",   7, ""),
        ("Wed 5/13", "Recovery + Flex",  "Weighted Pull-Up",    3, 8, "+5",   7, ""),
        ("Wed 5/13", "Recovery + Flex",  "Wtd Side Split Hold", 1, "60s", "10",  "-", ""),
        ("Wed 5/13", "Recovery + Flex",  "Wtd Side Split Hold", 2, "60s", "10",  "-", "stretch good"),
        ("Wed 5/13", "Recovery + Flex",  "Wtd Side Split Hold", 3, "60s", "10",  "-", ""),
        # Tue 5/12
        ("Tue 5/12", "Speed-Strength B", "Hang Power Snatch",   1, 3, "110",  7, ""),
        ("Tue 5/12", "Speed-Strength B", "Hang Power Snatch",   2, 3, "110",  7, ""),
        ("Tue 5/12", "Speed-Strength B", "Hang Power Snatch",   3, 3, "110",  7, ""),
        ("Tue 5/12", "Speed-Strength B", "Hang Power Snatch",   4, 3, "110",  8, ""),
        ("Tue 5/12", "Speed-Strength B", "Hang Power Snatch",   5, 3, "110",  8, ""),
        ("Tue 5/12", "Speed-Strength B", "DB Push Press",       1, 4, "45",   7, "shoulder fine"),
        # Mon 5/11
        ("Mon 5/11", "Upper Horizontal", "Bench Press",         1, 6, "170",  7, ""),
        ("Mon 5/11", "Upper Horizontal", "Bench Press",         2, 6, "170",  7, ""),
        ("Mon 5/11", "Upper Horizontal", "Bench Press",         3, 5, "170",  8, ""),
        ("Mon 5/11", "Upper Horizontal", "Bench Press",         4, 5, "170",  8, ""),
        ("Mon 5/11", "Upper Horizontal", "Barbell Row",         1, 6, "165",  7, ""),
        # Sun 5/10
        ("Sun 5/10", "Lower (LIGHT)",    "Back Squat",          1, 5, "280",  7, "top set"),
        ("Sun 5/10", "Lower (LIGHT)",    "Back Squat",          2, 5, "265",  7, ""),
        ("Sun 5/10", "Lower (LIGHT)",    "Back Squat",          3, 5, "265",  7, ""),
        ("Sun 5/10", "Lower (LIGHT)",    "Stiff-Legged DL",     1, 5, "195",  7, ""),
        ("Sun 5/10", "Lower (LIGHT)",    "Stiff-Legged DL",     2, 5, "195",  7, ""),
        ("Sun 5/10", "Lower (LIGHT)",    "Stiff-Legged DL",     3, 4, "195",  8, ""),
        ("Sun 5/10", "Lower (LIGHT)",    "Stiff-Legged DL",     4, 3, "195",  8, "heavy"),
    ]

    last_date = None
    for entry in logs:
        date, sess, ex, st, reps, load, rpe, note = entry
        # Banding by date
        fill = F_HEAD if date != last_date and last_date is not None else None
        # Actually band by alternating date
        ...
        last_date = date
        table_row(ws, row, [date, sess, ex, st, reps, load, rpe, note], align=L)
        row += 1

    # Add Excel autofilter
    last_col = get_column_letter(span)
    ws.auto_filter.ref = f"A4:{last_col}{row - 1}"
    # Freeze header
    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════════
#                            TAB 5 — PROGRESS
# ══════════════════════════════════════════════════════════════════════════
def build_progress(ws):
    widths(ws, 4, 20, 14, 14, 14, 14, 14, 14, 14)
    span = 9

    title_bar(
        ws, 1,
        "PROGRESS",
        "Bodyweight trend  ·  Top sets per arc week  ·  All-time PRs  ·  Weekly volume by category",
        span,
    )
    row = 4

    # ── BODYWEIGHT chart ──
    section(ws, row, "BODYWEIGHT — 7-day rolling avg vs target curve", span); row += 1
    # Data block (will be referenced by chart)
    bw_data_row = row
    ws.cell(row=row, column=2, value="Week").font = BODY_BOLD
    ws.cell(row=row, column=3, value="Actual (lb)").font = BODY_BOLD
    ws.cell(row=row, column=4, value="Target (lb)").font = BODY_BOLD
    row += 1
    bw_series = [
        ("Wk 1",  192.5, 192.0),
        ("Wk 2",  192.1, 191.3),
        ("Wk 3",  None,  190.7),
        ("Wk 4",  None,  190.0),
        ("Wk 5",  None,  189.3),
        ("Wk 6",  None,  188.0),
        ("Wk 7",  None,  187.5),
        ("Wk 8",  None,  187.0),
    ]
    for w, actual, target in bw_series:
        ws.cell(row=row, column=2, value=w).alignment = C
        ws.cell(row=row, column=3, value=actual).alignment = C
        ws.cell(row=row, column=4, value=target).alignment = C
        row += 1

    chart = LineChart()
    chart.title = "Bodyweight (7-day avg) — actual vs target"
    chart.style = 2
    chart.y_axis.title = "lb"
    chart.x_axis.title = "Arc week"
    chart.height = 8
    chart.width = 18
    data = Reference(ws, min_col=3, min_row=bw_data_row, max_col=4, max_row=bw_data_row + len(bw_series))
    cats = Reference(ws, min_col=2, min_row=bw_data_row + 1, max_row=bw_data_row + len(bw_series))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"F{bw_data_row}")
    row += 2

    # ── TOP SET — Key Lifts ──
    section(ws, row, "TOP SET — Key Lifts (this arc)", span); row += 1
    table_header(ws, row, ["Lift", "Wk 1", "Wk 2", "Wk 3", "Wk 4", "Wk 5", "Wk 6", "Trend"], start_col=2)
    row += 1
    top_sets = [
        ("Back Squat (Sun, light)", "280 × 5", "280 × 5", "—",      "—",      "—",      "—",      "→ hold (maint)"),
        ("Bench Press",             "165 × 6", "170 × 6", "—",      "—",      "—",      "—",      "↑ +5"),
        ("Power Clean",             "135 × 3", "140 × 3", "—",      "—",      "—",      "—",      "↑ +5"),
        ("Hang Snatch",             "105 × 3", "110 × 3", "—",      "—",      "—",      "—",      "↑ +5"),
        ("Speed Squat",             "205 × 2", "215 × 2", "—",      "—",      "—",      "—",      "↑ +10"),
        ("Weighted Pull-Up (Fri)",  "+10 × 5", "+15 × 5", "—",      "—",      "—",      "—",      "↑ +5"),
        ("Barbell Jump",            "110 × 4", "110 × 4", "—",      "—",      "—",      "—",      "→ hold"),
        ("CMJ (Sat measure)",       "28.5\"",  "—",       "—",      "—",      "—",      "—",      "baseline"),
    ]
    for entry in top_sets:
        table_row(ws, row, entry, start_col=2, bold_first=True, align=C)
        row += 1
    row += 1

    # ── PRs ──
    section(ws, row, "PRs (all-time, auto-computed from Log)", span); row += 1
    table_header(ws, row, ["Lift", "Best", "Date", "Note"], start_col=2)
    row += 1
    prs = [
        ("Back Squat",        "370 × 1",  "Pre-arc",          "Baseline (1RM)"),
        ("Bench Press",       "230 × 1",  "Pre-arc",          "Baseline (1RM)"),
        ("Power Clean",       "140 × 3",  "Thu  May 14, 2026", "← ARC PR (vs 135 Wk 1)"),
        ("Weighted Pull-Up",  "+15 × 5",  "Fri  May 08, 2026", "← ARC PR (vs +10)"),
        ("Hang Snatch",       "110 × 3",  "Tue  May 12, 2026", "← ARC PR (vs 105 Wk 1)"),
        ("CMJ",               "28.5\"",   "Sat  May 09, 2026", "Arc-start baseline"),
    ]
    for entry in prs:
        fill = F_PR if "ARC PR" in entry[3] else None
        table_row(ws, row, entry, start_col=2, fill=fill, bold_first=True, align=L)
        row += 1
    row += 1

    # ── WEEKLY VOLUME chart ──
    section(ws, row, "WEEKLY VOLUME — Wk 2 (sets per movement category)", span); row += 1
    vol_data_row = row
    ws.cell(row=row, column=2, value="Category").font = BODY_BOLD
    ws.cell(row=row, column=3, value="Sets").font = BODY_BOLD
    row += 1
    volume = [
        ("Press",     16),
        ("Pull",      14),
        ("Squat",     12),
        ("Hinge",     10),
        ("Olympic",    9),
        ("Plyo",      11),
        ("Sprint",     9),
        ("Mobility", 11),
    ]
    for cat, n in volume:
        ws.cell(row=row, column=2, value=cat).alignment = L
        ws.cell(row=row, column=3, value=n).alignment = C
        row += 1

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Weekly volume — sets per movement category (Wk 2)"
    bar.style = 2
    bar.height = 9
    bar.width = 18
    data = Reference(ws, min_col=3, min_row=vol_data_row, max_row=vol_data_row + len(volume))
    cats = Reference(ws, min_col=2, min_row=vol_data_row + 1, max_row=vol_data_row + len(volume))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, f"F{vol_data_row}")


# ══════════════════════════════════════════════════════════════════════════
#                                  MAIN
# ══════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    # Rename + create tabs in order
    home = wb.active
    home.title = "Home"
    this_week = wb.create_sheet("This Week")
    plan = wb.create_sheet("Plan")
    log = wb.create_sheet("Log")
    progress = wb.create_sheet("Progress")

    build_home(home)
    build_this_week(this_week)
    build_plan(plan)
    build_log(log)
    build_progress(progress)

    # Make Home the landing tab
    wb.active = 0

    wb.save(OUT)
    print(f"✓ wrote {OUT}")
    print(f"  tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
