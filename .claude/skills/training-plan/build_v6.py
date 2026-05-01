"""
Generate hybrid-athletic-plan-v6.xlsx — Andy's 18-week dunk + upper + side split arc.

Power/reactive emphasis throughout (athlete is strength-dominant, has saturated on
strength blocks). VJ guide drives jump methodology; Dylan Shannon drives upper days
and 4-pillar lower distribution. 6-day week: 1 strength + 2 speed-strength + 1 jump
+ 2 upper + Wed optional flex.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============== Constants ==============
SQUAT_1RM = 370
BENCH_1RM = 230
CLEAN_1RM = 195

# ============== Style helpers ==============
FT = Font(name='Arial', size=16, bold=True, color='FFFFFF')
FH2 = Font(name='Arial', size=13, bold=True)
FH3 = Font(name='Arial', size=11, bold=True, color='FFFFFF')
FH4 = Font(name='Arial', size=11, bold=True)
FB = Font(name='Arial', size=10)
FB_BOLD = Font(name='Arial', size=10, bold=True)
FB_ITAL = Font(name='Arial', size=10, italic=True)

DARK = PatternFill('solid', fgColor='1A1A2E')
SECTION = PatternFill('solid', fgColor='F0F4FF')
GOAL = PatternFill('solid', fgColor='E6F4EA')
WARN = PatternFill('solid', fgColor='FFF3E0')
B1_FILL = PatternFill('solid', fgColor='DBEAFE')
B2_FILL = PatternFill('solid', fgColor='FDE68A')
B3_FILL = PatternFill('solid', fgColor='FECACA')
FLEX = PatternFill('solid', fgColor='F3E8FF')
SUPER = PatternFill('solid', fgColor='F1F5F9')
PROFILE = PatternFill('solid', fgColor='EFF6FF')

ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def pct_wt(pct, max_lb):
    w = round((pct / 100) * max_lb / 5) * 5
    return f"{pct}% ({w} lb)"


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write(ws, coord, value, font=None, fill=None, align=None):
    cell = ws[coord]
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align


def title_row(ws, row, text, fill=DARK):
    """Span A:I (or whatever max) with title."""
    ws.row_dimensions[row].height = 28
    write(ws, f'A{row}', text, font=FT, fill=fill, align=ALIGN_LEFT)
    for col in range(2, 10):
        ws.cell(row=row, column=col).fill = fill


def section_header(ws, row, text, fill=SECTION):
    ws.row_dimensions[row].height = 22
    write(ws, f'A{row}', text, font=FH2, fill=fill, align=ALIGN_LEFT)
    for col in range(2, 10):
        ws.cell(row=row, column=col).fill = fill


def day_header(ws, row, text):
    ws.row_dimensions[row].height = 22
    write(ws, f'A{row}', text, font=FH3, fill=DARK, align=ALIGN_LEFT)
    for col in range(2, 10):
        ws.cell(row=row, column=col).fill = DARK


def grid_header(ws, row):
    """# | Exercise | Sets×Reps | Wk1..Wk6"""
    headers = ['#', 'Exercise', 'Sets × Reps', 'Wk 1', 'Wk 2', 'Wk 3', 'Wk 4', 'Wk 5', 'Wk 6']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i)
        c.value = h
        c.font = FH4
        c.fill = SECTION
        c.alignment = ALIGN_CENTER


def grid_row(ws, row, num, exercise, setsreps, week_loads, fill=None, font=None):
    """Write one exercise row.  week_loads = list of 6 strings."""
    vals = [num, exercise, setsreps] + list(week_loads)
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i)
        c.value = v
        c.font = font or FB
        if fill:
            c.fill = fill
        c.alignment = ALIGN_LEFT if i == 2 else ALIGN_CENTER


# ============== Build workbook ==============
wb = Workbook()

# ===== Athlete Profile =====
ws = wb.active
ws.title = 'Athlete Profile'
set_widths(ws, [22, 28, 28, 36])

title_row(ws, 1, 'ATHLETE PROFILE — ANDY LEE', fill=PatternFill('solid', fgColor='1E3A8A'))

section_header(ws, 3, 'IDENTITY', fill=PROFILE)
profile_rows = [
    ('Name', 'Andy Lee', '', ''),
    ('Location', 'Los Angeles', '', ''),
    ('Training Age', '~1.5 years serious', '', ''),
    ('Identity', 'Athlete for life — basketball, martial arts, strength', '', ''),
    ('Profile', 'Strength-dominant, reactivity-deficient (per VJ §3 dx)', '1.93× BW squat', 'Bottleneck = RFD, not max force'),
]
for i, r in enumerate(profile_rows, 4):
    write(ws, f'A{i}', r[0], font=FB_BOLD)
    write(ws, f'B{i}', r[1], font=FB)
    write(ws, f'C{i}', r[2], font=FB)
    write(ws, f'D{i}', r[3], font=FB_ITAL)

section_header(ws, 10, 'CURRENT STATS')
write(ws, 'A11', 'Metric', font=FH4); write(ws, 'B11', 'Current', font=FH4)
write(ws, 'C11', 'Target', font=FH4); write(ws, 'D11', 'Notes', font=FH4)
stats = [
    ('Bodyweight', '~192 lb', '183-186 lb', 'Cut ~300 cal/day Blocks 1-2; maintenance Block 3'),
    ('Body Fat', '~15-16%', '~11%', 'Power-to-weight ratio is jump multiplier'),
    ('Back Squat', '~370 lb (1.93× BW)', 'Maintain ≥ 360', 'Ceiling already high; reactivity is the bottleneck'),
    ('Bench Press', '230 lb', '260 lb', 'Arc G2 — Dylan-style volume push'),
    ('Power Clean', '~195 lb est', '210+ lb', 'Speed-strength expression'),
    ('Hang Power Snatch', '~140 lb est', '160+ lb', 'Saturday variety — technique focus'),
    ('Weighted Pull-Up', '+25 lb × 5', '+80 lb × 8', 'Arc G2 — peaked Block 3'),
    ('Standing CMJ', 'TBD baseline', '+4 in by Wk 18', 'Reactive deficit test (VJ §3)'),
    ('Approach Touch', 'Can grab rim', 'Above rim → DUNK women\'s BB', 'Arc G1 — film weekly Block 2-3'),
    ('Side Split (cm to floor)', 'TBD baseline', '−8 cm cumulative', 'Arc G3 — measure weekly'),
]
for i, r in enumerate(stats, 12):
    write(ws, f'A{i}', r[0], font=FB_BOLD)
    write(ws, f'B{i}', r[1], font=FB)
    write(ws, f'C{i}', r[2], font=FB)
    write(ws, f'D{i}', r[3], font=FB_ITAL)

start_row = 12 + len(stats) + 1
section_header(ws, start_row, 'ACTIVE INJURIES', fill=WARN)
inj_rows = [
    ('Right Shoulder', 'Active, manageable', 'No barbell OHP, no behind-neck', 'Landmine, neutral DB, floor press'),
    ('Left Wrist (De Quervain\'s)', 'Active', 'No front rack clean grip', 'Hang variants only; straps + hook grip on pulls'),
]
for i, r in enumerate(inj_rows, start_row + 1):
    write(ws, f'A{i}', r[0], font=FB_BOLD, fill=WARN)
    write(ws, f'B{i}', r[1], font=FB, fill=WARN)
    write(ws, f'C{i}', r[2], font=FB, fill=WARN)
    write(ws, f'D{i}', r[3], font=FB_ITAL, fill=WARN)

start_row += len(inj_rows) + 2
section_header(ws, start_row, 'SCHEDULE + CONSTRAINTS')
sched_rows = [
    ('Training Days', 'Sun, Mon, Tue, Thu, Fri, Sat', 'Wed = optional flex/recovery', '6-day week (Dylan template)'),
    ('Session Length', '60-75 min (Sat 45 min)', 'Max 6 exercises', 'Daily flex 15-20 min not counted'),
    ('Equipment', 'Full commercial gym + sprint surface', 'Barbell, rack, platform, cables, sled', ''),
    ('Nutrition', '~300 cal deficit (Blocks 1-2)', 'Maintenance Block 3 (peak)', 'Pause cut if squat drops >5%'),
]
for i, r in enumerate(sched_rows, start_row + 1):
    write(ws, f'A{i}', r[0], font=FB_BOLD)
    write(ws, f'B{i}', r[1], font=FB)
    write(ws, f'C{i}', r[2], font=FB)
    write(ws, f'D{i}', r[3], font=FB_ITAL)

start_row += len(sched_rows) + 2
section_header(ws, start_row, 'TRAINING PREFERENCES')
prefs = [
    ('Fixed exercise menu per block', 'Progress through load, not variation'),
    ('Supersets always', 'Except power/Olympic work (always standalone)'),
    ('RPE-based intensity', 'RPE 6-8 most work; no grinders except test weeks'),
    ('Hardest day first', 'Sun = heaviest squat; Tue = jump peak day'),
    ('Daily flex protocol', '15-20 min post-training (deep work Wed only)'),
    ('Style emphasis', 'VJ (primary, jump methodology) + Dylan (secondary, upper + 4 pillars)'),
]
for i, p in enumerate(prefs, start_row + 1):
    write(ws, f'A{i}', p[0], font=FB_BOLD)
    write(ws, f'B{i}', p[1], font=FB_ITAL)

# ===== Arc =====
ws = wb.create_sheet('Arc')
set_widths(ws, [10, 30, 14, 26, 18, 14, 14, 14])

title_row(ws, 1, 'HYBRID ATHLETIC ARC — 18 WEEKS · DUNK + UPPER + SPLIT')

section_header(ws, 3, 'PURPOSE')
write(ws, 'A4',
      'Close the explosive strength deficit so a strength-dominant athlete can finally express '
      'his force as a dunk. Build real upper body strength on the side. Drive the side split as a '
      'daily protocol. VJ drives the jump methodology; Dylan drives the upper days and the 4-pillar '
      'lower distribution.', font=FB)
ws.merge_cells('A4:H4')
ws.row_dimensions[4].height = 50

section_header(ws, 6, 'GOALS')
write(ws, 'A7', 'ID', font=FH4); write(ws, 'B7', 'Goal', font=FH4)
write(ws, 'C7', 'Yes/No?', font=FH4); write(ws, 'D7', 'Test Method', font=FH4)
write(ws, 'E7', 'Deadline', font=FH4)
arc_goals = [
    ('G1', 'Dunk a women\'s basketball on a regulation 10\' rim', 'Yes', 'Filmed clean dunk attempt', 'Wk 18'),
    ('G2', 'Bench 1RM 260 lb AND weighted pull-up +80 lb × 8 clean', 'Yes', '1RM test + max-reps pull-up', 'Wk 18'),
    ('G3', 'Side split distance reduced ≥ 8 cm from baseline (hip-to-floor)', 'Yes', 'Tape measure, weekly', 'Wk 18'),
]
for i, g in enumerate(arc_goals, 8):
    for j, v in enumerate(g, 1):
        c = ws.cell(row=i, column=j)
        c.value = v
        c.font = FB_BOLD if j <= 2 else FB
        c.fill = GOAL
        c.alignment = ALIGN_LEFT if j == 2 else ALIGN_CENTER

section_header(ws, 12, 'BLOCK SEQUENCE')
write(ws, 'A13', 'Block', font=FH4); write(ws, 'B13', 'Name', font=FH4)
write(ws, 'C13', 'Weeks', font=FH4); write(ws, 'D13', 'Purpose', font=FH4)
write(ws, 'E13', 'Serves', font=FH4)
blocks = [
    ('1', 'Power Conversion + Upper Build', '1–6',
     'Close ESD: jump squats + Olympic + reactive intro; squat on maintenance. Bench/pull-up volume base.', 'G1, G2, G3'),
    ('2', 'Reactive + Dunk Window', '7–12',
     'Depth jumps enter; dunk attempts begin Wk 9. Bench peaks toward 255; pull-up toward +50.', 'G1, G2, G3'),
    ('3', 'Peak + Realize + Test', '13–18',
     'Volume taper; PAP complexes; dunk attempts every Tue. Test bench 260 + pull-up +80×8 + dunk.', 'G1, G2, G3'),
]
fills = [B1_FILL, B2_FILL, B3_FILL]
for i, b in enumerate(blocks, 14):
    for j, v in enumerate(b, 1):
        c = ws.cell(row=i, column=j)
        c.value = v
        c.font = FB_BOLD if j <= 1 else FB
        c.fill = fills[i - 14]
        c.alignment = ALIGN_LEFT if j in (2, 4) else ALIGN_CENTER

section_header(ws, 18, 'TESTING SCHEDULE')
write(ws, 'A19', 'Metric', font=FH4); write(ws, 'B19', 'Baseline', font=FH4)
write(ws, 'C19', 'Wk 6', font=FH4); write(ws, 'D19', 'Wk 12', font=FH4)
write(ws, 'E19', 'Wk 18', font=FH4); write(ws, 'F19', 'Method', font=FH4)
write(ws, 'G19', 'Serves', font=FH4)
tests = [
    ('Standing CMJ', 'TBD', '+1 in', '+2.5 in', '+4 in', 'Best of 3, jump mat or wall', 'G1'),
    ('Bounce Depth Jump (18 in)', 'TBD', '≥ Standing CMJ', '> Standing', 'Reactive index closed', 'Best of 3 from box', 'G1'),
    ('Approach Touch', 'Grab rim', 'Consistent grab', '+3 in above rim', 'DUNK', 'Film + wall mark', 'G1'),
    ('Back Squat', '~370 lb', '≥ 360 (maintained)', '≥ 360', '≥ 360', 'Top set 3RM', 'G1 ceiling'),
    ('Bench Press', '230 lb', '245 (3RM)', '255 (1RM)', '260 (1RM)', '1RM or e1RM', 'G2'),
    ('Wtd Pull-Up', '+25 lb × 5', '+25 × 5 clean', '+50 × 5', '+80 × 8', 'Max reps @ target', 'G2'),
    ('Side Split', 'TBD baseline', '−3 cm', '−6 cm', '−8 cm', 'Tape, hip-to-floor, weekly', 'G3'),
    ('Bodyweight', '~192 lb', '~188', '~185', '183-186 (held)', 'Daily, 7-day avg', 'G1 multiplier'),
]
for i, t in enumerate(tests, 20):
    for j, v in enumerate(t, 1):
        c = ws.cell(row=i, column=j)
        c.value = v
        c.font = FB_BOLD if j == 1 else FB
        c.alignment = ALIGN_LEFT if j == 1 else ALIGN_CENTER

next_row = 20 + len(tests) + 1
section_header(ws, next_row, 'CONSTRAINTS', fill=WARN)
constraints = [
    'R shoulder: no barbell OHP — landmine/DB neutral grip only',
    'L wrist: no front rack — hang Olympic variants + hook grip + straps on pulls',
    'Cut ~300 cal/day Blocks 1-2 only; maintenance calories Block 3 (peak)',
    'Side split protocol post-session only on Sun/Tue/Thu (loaded splits sap jump quality 24-48h)',
    'Deep split work = Wed only (true rest day for legs)',
    'If squat top set drops >5% in a week → halve calorie deficit; if >8% → pause cut',
    'Patellar pain ≥ 3/10 → drop depth jumps that week; ≥ 5 → 2-wk plyo deload',
]
for i, c in enumerate(constraints, next_row + 1):
    write(ws, f'A{i}', '•', font=FB_BOLD, fill=WARN)
    write(ws, f'B{i}', c, font=FB, fill=WARN, align=ALIGN_LEFT)
    ws.merge_cells(f'B{i}:H{i}')
    for col in range(3, 9):
        ws.cell(row=i, column=col).fill = WARN

# ============== Block tabs ==============
def build_block(name, block_num, fill, purpose, goals, weekly_split, prog_model, days, display_title=None):
    """days = list of (day_name, day_title, exercises) where exercises is list of dicts."""
    ws = wb.create_sheet(name)
    set_widths(ws, [6, 32, 16, 17, 17, 17, 17, 17, 17])

    title_text = f'BLOCK {block_num}: {display_title}' if display_title else f'BLOCK {block_num}'
    title_row(ws, 1, title_text, fill=DARK)
    for col in range(1, 10):
        ws.cell(row=1, column=col).fill = fill
        ws.cell(row=1, column=col).font = Font(name='Arial', size=16, bold=True)

    section_header(ws, 3, 'PURPOSE')
    write(ws, 'A4', purpose, font=FB, align=ALIGN_LEFT)
    ws.merge_cells('A4:I4')
    ws.row_dimensions[4].height = 45

    # Goals
    section_header(ws, 6, 'GOALS')
    for i, g in enumerate(goals, 7):
        write(ws, f'A{i}', g[0], font=FB_BOLD, fill=GOAL)
        write(ws, f'B{i}', g[1], font=FB, fill=GOAL, align=ALIGN_LEFT)
        ws.merge_cells(f'B{i}:G{i}')
        for col in range(3, 8):
            ws.cell(row=i, column=col).fill = GOAL
        write(ws, f'H{i}', g[2], font=FB_ITAL, fill=GOAL, align=ALIGN_CENTER)
        ws.merge_cells(f'H{i}:I{i}')
        for col in range(8, 10):
            ws.cell(row=i, column=col).fill = GOAL

    next_row = 7 + len(goals) + 1
    section_header(ws, next_row, 'WEEKLY SPLIT')
    write(ws, f'A{next_row + 1}', 'Day', font=FH4)
    write(ws, f'B{next_row + 1}', 'Workout', font=FH4)
    for i, (d, w) in enumerate(weekly_split, next_row + 2):
        write(ws, f'A{i}', d, font=FB_BOLD)
        write(ws, f'B{i}', w, font=FB, align=ALIGN_LEFT)

    next_row = next_row + 2 + len(weekly_split) + 1
    section_header(ws, next_row, 'PROGRESSION MODEL')
    for i, line in enumerate(prog_model, next_row + 1):
        write(ws, f'A{i}', line, font=FB, align=ALIGN_LEFT)
        ws.merge_cells(f'A{i}:I{i}')

    next_row = next_row + 1 + len(prog_model) + 1
    section_header(ws, next_row, 'DAILY BREAKDOWN')

    cur_row = next_row + 2
    for day_name, day_title, exercises in days:
        day_header(ws, cur_row, f'{day_name} — {day_title}')
        cur_row += 1
        grid_header(ws, cur_row)
        cur_row += 1
        for ex in exercises:
            row_fill = None
            if ex.get('flex'):
                row_fill = FLEX
            elif ex.get('superset_b'):
                row_fill = SUPER
            grid_row(ws, cur_row, ex['num'], ex['exercise'], ex['setsreps'], ex['weeks'], fill=row_fill)
            cur_row += 1
        cur_row += 1  # spacer between days


# ============== BLOCK 1 ==============
b1_purpose = (
    'Close the explosive strength deficit. Convert existing strength into expressed power. '
    'Squat moves to maintenance (1×/wk Sun); jump squats / Olympic lifts / sprints get the volume. '
    'Build real upper body base for bench + pull-up. Establish the daily flex protocol.'
)
b1_goals = [
    ('G1', 'Standing CMJ +1 inch from baseline AND barbell jump 4×4 @ 33% squat (120 lb) clean', 'Serves: Arc G1'),
    ('G2', 'Bench 3RM ≥ 245 lb (from 230) AND weighted pull-up +25 lb × 5 reps clean both sessions', 'Serves: Arc G2'),
    ('G3', 'Side split distance reduced ≥ 3 cm from baseline (1 deep day/wk = conservative)', 'Serves: Arc G3'),
]
b1_split = [
    ('Sun', 'Lower Strength (LIGHT maintenance) + Posterior Chain'),
    ('Mon', 'Upper Horizontal — Bench + Row'),
    ('Tue', 'Speed-Strength B — Olympic Variety + Sprint Finisher'),
    ('Wed', 'Recovery + Deep Flex (Optional) + Pull-Up Volume'),
    ('Thu', 'Speed-Strength A — Olympic Power + Speed Squat'),
    ('Fri', 'Upper Vertical — Pull-Up + Dips'),
    ('Sat', 'JUMP LAB — Sprints + Approach + Depth + Loaded Power'),
]
b1_progression = [
    'WEEK SHAPE: Sat = peak performance day (jump). Thu = power buildup. Sun = lightest squat (must NOT compete with Sat 24h later).',
    'Squat: TRUE MAINTENANCE — Sun top set 75→80% (275→295 lb), 1×5 + 2×5 back-off lighter. Just keeps the pattern warm; never push.',
    'Bench: +2%/wk wave (72→82%, 165→190 lb). Push volume — 4×6 → 4×5 as load climbs. Test 3RM Wk 6.',
    'Pull-Up: Fri 4×5 progresses +5 lb/wk (+10 → +25). Tue 4×5 trails Fri by ~5 lb. Wed 3×8 volume at lighter load.',
    'Power Clean (Thu): 5×3 wave 70→78% (135→155 lb). Hang Power Clean (Sat) 4×3 at 65→75%.',
    'Hang Snatch (Tue): 5×3 wave 55→65% (105→125 lb). DB Push Press neutral 4×4.',
    'Speed Squat (Thu): 5×2 wave 55→65% squat 1RM. Jump Squat (Thu) 4×4 @ 25→30%. Barbell Jump (Sat) 4×4 @ 30→33%.',
    'Sprints: Sat 3×20yd + 2×30yd full effort (jump day); Tue 3×20yd + 1×40yd finisher (variety day).',
    'Reactive: Low-box depth jump (18 in) Sat 2×5 from Wk 1 — early intro since athlete is reactivity-deficient.',
    'SLDL/RDL: wave-loaded 4×5,5,4,3, +10 lb/wk (185→225). Same weight all 4 sets within session.',
    'Side split: Wed deep work — weighted hold 3×60s, +5 lb every 2 wks (10→20 lb). PNF C-R 3 rounds.',
    'CNS spacing: Sat jump → Sun light squat 24h (light only) → Mon upper → Tue speed-strength 48h post squat ✓ → Thu power 48h post Tue ✓ → Sat jump 48h post Thu ✓.',
    'Wk 1-2 RPE 6-7. Wk 3-4 RPE 7-8. Wk 5 RPE 8-9 (peak). Wk 6 deload (60% volume, drop intensity 10%).',
]

# Block 1 daily exercises
def ex(num, exercise, setsreps, weeks, **kw):
    return {'num': num, 'exercise': exercise, 'setsreps': setsreps, 'weeks': weeks, **kw}

b1_sun = [
    ex('1a', 'Back Squat (light maintenance)', '1×5 + 2×5', [pct_wt(75, SQUAT_1RM), pct_wt(75, SQUAT_1RM),
                                                              pct_wt(78, SQUAT_1RM), pct_wt(78, SQUAT_1RM),
                                                              pct_wt(80, SQUAT_1RM), pct_wt(70, SQUAT_1RM)]),
    ex('2a', 'Stiff-Legged Deadlift', '4×5,5,4,3', ['185 lb', '195', '205', '215', '225', '175']),
    ex('3a', 'Hand-Supported Bulgarian SS', '3×8/leg', ['30 lb DB', '35', '35', '40', '45', '25']),
    ex('4a', 'Glute-Ham Raise', '3×8', ['BW', 'BW', 'BW+5', 'BW+5', 'BW+10', 'BW']),
    ex('5a', 'Pigeon Stretch', '2×60s/side', ['—'] * 6, flex=True),
]
b1_mon = [
    ex('1a', 'Bench Press', '4×6 → 4×5', [pct_wt(72, BENCH_1RM), pct_wt(74, BENCH_1RM), pct_wt(77, BENCH_1RM),
                                            pct_wt(80, BENCH_1RM), pct_wt(82, BENCH_1RM), pct_wt(70, BENCH_1RM)]),
    ex('2a', 'Barbell Row', '4×6', ['155 lb', '165', '170', '175', '180', '145']),
    ex('3a', 'DB Incline Bench (neutral)', '3×6-8', ['50 lb', '55', '55', '60', '60', '45']),
    ex('3b', 'T-Bar Chest-Supported Row', '3×6-8', ['Mod', 'Mod', 'Mod-Hvy', 'Heavy', 'Heavy', 'Mod'], superset_b=True),
    ex('4a', 'Cable Fly', '3×8-12', ['Mod', 'Mod', 'Mod', 'Mod', 'Mod', 'Light']),
    ex('4b', 'DB Lateral Raise', '3×10', ['15 lb', '15', '17.5', '17.5', '20', '12.5'], superset_b=True),
]
b1_tue = [
    ex('1a', 'Sprints', '3×20yd, 2×30yd', ['95% effort', '95%', '95%', '95%', '95%', '2×20yd']),
    ex('2a', 'Approach Jumps', '5×2', ['max intent', 'max', 'max', 'max', 'max', '3×2 deload']),
    ex('3a', 'Low-Box Depth Jump', '2×5', ['18 in', '18 in', '18 in', '18 in', '18 in', 'omit']),
    ex('4a', 'Barbell Jump', '4×4', [pct_wt(30, SQUAT_1RM), pct_wt(30, SQUAT_1RM), pct_wt(30, SQUAT_1RM),
                                       pct_wt(33, SQUAT_1RM), pct_wt(33, SQUAT_1RM), pct_wt(25, SQUAT_1RM)]),
    ex('5a', 'Hang Power Clean', '4×3', [pct_wt(65, CLEAN_1RM), pct_wt(68, CLEAN_1RM), pct_wt(70, CLEAN_1RM),
                                           pct_wt(72, CLEAN_1RM), pct_wt(75, CLEAN_1RM), pct_wt(60, CLEAN_1RM)]),
]
b1_wed = [
    ex('1a', 'Weighted Pull-Up (volume)', '3×8', ['+5 lb', '+5', '+10', '+10', '+15', '+5']),
    ex('2a', 'Face Pull', '3×15', ['Light', 'Light', 'Light', 'Light', 'Light', 'Light']),
    ex('3a', 'Weighted Side Split Hold', '3×60s', ['10 lb', '10', '15', '15', '20', '10'], flex=True),
    ex('3b', 'PNF Contract-Relax Splits', '3 rounds', ['10s on/30s relax'] * 6, flex=True, superset_b=True),
    ex('4a', 'Pigeon Stretch', '2×60s/side', ['—'] * 6, flex=True),
    ex('4b', 'Cossack Squat', '3×8/side', ['BW'] * 6, flex=True, superset_b=True),
]
b1_thu = [
    ex('1a', 'Power Clean', '5×3', [pct_wt(70, CLEAN_1RM), pct_wt(72, CLEAN_1RM), pct_wt(75, CLEAN_1RM),
                                      pct_wt(75, CLEAN_1RM), pct_wt(78, CLEAN_1RM), pct_wt(65, CLEAN_1RM)]),
    ex('2a', 'Speed Squat', '5×2', [pct_wt(55, SQUAT_1RM), pct_wt(58, SQUAT_1RM), pct_wt(60, SQUAT_1RM),
                                      pct_wt(62, SQUAT_1RM), pct_wt(65, SQUAT_1RM), pct_wt(55, SQUAT_1RM)]),
    ex('3a', 'Jump Squat', '4×4', [pct_wt(25, SQUAT_1RM), pct_wt(25, SQUAT_1RM), pct_wt(28, SQUAT_1RM),
                                     pct_wt(28, SQUAT_1RM), pct_wt(30, SQUAT_1RM), pct_wt(20, SQUAT_1RM)]),
    ex('4a', 'RDL', '4×5,5,4,3', ['185 lb', '195', '205', '215', '225', '175']),
    ex('5a', 'Pendulum Squat', '3×8-10', ['Mod', 'Mod', 'Mod-Hvy', 'Heavy', 'Heavy', 'Mod']),
]
b1_fri = [
    ex('1a', 'Weighted Pull-Up', '4×5', ['+10 lb', '+15', '+20', '+25', '+25', '+10']),
    ex('2a', 'Weighted Dips', '3×8', ['+10 lb', '+15', '+15', '+20', '+20', '+10']),
    ex('3a', 'Chest-Supported Row', '4×6-8', ['Mod', 'Mod', 'Mod-Hvy', 'Heavy', 'Heavy', 'Mod']),
    ex('3b', 'Landmine Press', '3×10', ['Mod', 'Mod', 'Mod', 'Mod-Hvy', 'Mod-Hvy', 'Light'], superset_b=True),
    ex('4a', 'Incline DB Curl', '3×8-10', ['Mod', 'Mod', 'Mod', 'Mod', 'Mod', 'Light']),
    ex('4b', 'Tricep Pushdown', '3×8-10', ['Mod', 'Mod', 'Mod', 'Mod', 'Mod', 'Light'], superset_b=True),
]
b1_sat = [
    ex('1a', 'Hang Power Snatch', '5×3', ['55% (105 lb)', '58% (110)', '60% (115)', '62% (120)', '65% (125)', '50% (95)']),
    ex('2a', 'DB Push Press (neutral)', '4×4', ['40 lb DB', '45', '45', '50', '50', '35']),
    ex('3a', 'Single-Leg Pogos', '3×8/leg', ['BW'] * 6),
    ex('4a', 'Sprints (finisher)', '3×20yd, 1×40yd', ['90% effort', '90%', '90%', '90%', '90%', '2×20yd']),
    ex('5a', 'Copenhagen Lift', '3×15s/side', ['BW'] * 6),
]

build_block('Block 1', 1, B1_FILL, b1_purpose, b1_goals, b1_split, b1_progression, [
    ('SUNDAY', 'Lower Strength (LIGHT maintenance) + Posterior', b1_sun),
    ('MONDAY', 'Upper Horizontal — Bench + Row', b1_mon),
    ('TUESDAY', 'Speed-Strength B — Olympic Variety + Sprint Finisher', b1_sat),
    ('WEDNESDAY', 'Recovery + Deep Flex (Optional) + Pull-Up Volume', b1_wed),
    ('THURSDAY', 'Speed-Strength A — Olympic Power + Speed Squat', b1_thu),
    ('FRIDAY', 'Upper Vertical — Pull-Up + Dips', b1_fri),
    ('SATURDAY', 'JUMP LAB — Sprint + Approach + Depth + Loaded Power', b1_tue),
], display_title='POWER CONVERSION + UPPER BUILD')

# ============== BLOCK 2 ==============
b2_purpose = (
    'Reactive emphasis: depth jumps enter the program. Dunk attempts begin Wk 9. Bench '
    'and pull-up push toward arc targets — Mon adds density (Dylan-style 14-min clock @ 80% Wks 9-11). '
    'Squat stays maintenance with PAP complexes introduced (heavy → explosive sequencing).'
)
b2_goals = [
    ('G1', 'Approach touch ≥ 3 inches above rim AND first dunk attempts filmed (Wk 9+)', 'Serves: Arc G1'),
    ('G2', 'Bench 1RM ≥ 255 lb AND weighted pull-up +50 lb × 5 reps clean', 'Serves: Arc G2'),
    ('G3', 'Side split distance reduced ≥ 6 cm cumulative from baseline', 'Serves: Arc G3'),
]
b2_split = [
    ('Sun', 'Lower Strength (LIGHT maintenance) + Posterior Chain'),
    ('Mon', 'Upper Horizontal — Bench Density + Row'),
    ('Tue', 'Speed-Strength B — Olympic Variety + Sprint Finisher'),
    ('Wed', 'Recovery + Deep Flex + Pull-Up Volume'),
    ('Thu', 'Speed-Strength A — Olympic Power + Loaded Velocity'),
    ('Fri', 'Upper Vertical — Pull-Up Heavy + Dips'),
    ('Sat', 'JUMP LAB + Dunk Attempts (Wk 9+) — Mid/High-Box Depth Jumps'),
]
b2_progression = [
    'WEEK SHAPE: Sat = peak performance (jump + dunk attempts). Sun stays LIGHT maintenance (only 24h after Sat jump).',
    'Sun: TRUE MAINTENANCE squat 3×3 @ 78-80%. NO PAP complex (would compete with Sat reactive work). Posterior chain holds heavy.',
    'Bench Mon: Wks 7-8 conventional 5×5 climbing 185→195. Wks 9-11 14-min density clock @ 80% (185 lb), AMRAP-style total reps. Wk 12 test 1RM (target 255).',
    'Pull-Up Fri: heavy triples climb 5×3 from +30→+50. Tue 5×3 trails by ~5 lb. Wed 3×8 volume at 60% Fri load.',
    'Power Clean Thu: 5×3 wave 75→82% (145→160 lb). Hang Power Clean Sat 4×3 at 72→78%.',
    'Hang Snatch Tue: 5×3 wave 60→70% (115→135 lb). DB Push Press 4×4 climbing.',
    'Depth Jump (Sat): Wk 7-8 mid box (24 in) 3×5; Wk 9-11 high box (30 in) 3×5; Wk 12 deload 18 in 2×5. Max intent every contact.',
    'Dunk attempts begin Wk 9: 5-10 max-intent attempts at end of SATURDAY session, after the depth jumps. Film every attempt.',
    'Speed Squat Thu: 5×2 wave 65→72% (240→265 lb). Jump Squat 4×4 @ 30→35%.',
    'Sprints: Sat 4×20yd + 2×30yd (jump day); Tue 4×20yd + 1×40yd finisher.',
    'RDL/SLDL: wave continues +10 lb/wk from B1 endpoint. 235→275 lb.',
    'Wk 7-8 RPE 7-8 (build). Wk 9-11 RPE 8-9 (peak). Wk 12 deload (60% volume).',
]

b2_sun = [
    ex('1a', 'Back Squat (light maintenance)', '3×3', [pct_wt(78, SQUAT_1RM), pct_wt(78, SQUAT_1RM),
                                                         pct_wt(80, SQUAT_1RM), pct_wt(80, SQUAT_1RM),
                                                         pct_wt(80, SQUAT_1RM), pct_wt(72, SQUAT_1RM)]),
    ex('2a', 'Stiff-Legged Deadlift', '4×5,5,4,3', ['235 lb', '245', '255', '265', '275', '215']),
    ex('3a', 'Hand-Supported Bulgarian SS', '3×8/leg', ['45 lb', '50', '50', '55', '55', '40']),
    ex('4a', 'Glute-Ham Raise', '3×8', ['BW+10', 'BW+10', 'BW+15', 'BW+15', 'BW+20', 'BW+10']),
    ex('5a', 'Pigeon Stretch', '2×60s/side', ['—'] * 6, flex=True),
]
b2_mon = [
    ex('1a', 'Bench Press', 'Wk7-8: 5×5 / Wk9-11: 14-min density / Wk12: test',
       ['185 lb 5×5', '195 5×5', '14-min @ 185', '14-min @ 190', '14-min @ 195', 'Test 1RM']),
    ex('2a', 'Barbell Row', '4×6', ['185 lb', '190', '195', '200', '205', '170']),
    ex('3a', 'DB Incline Bench (neutral)', '3×6-8', ['60 lb', '65', '65', '70', '70', '50']),
    ex('3b', 'T-Bar Chest-Supported Row', '3×6-8', ['Heavy']*5 + ['Mod'], superset_b=True),
    ex('4a', 'Cable Fly', '3×8-12', ['Mod-Hvy']*5 + ['Light']),
    ex('4b', 'DB Lateral Raise', '3×10', ['20 lb', '20', '22.5', '22.5', '25', '15'], superset_b=True),
]
b2_tue = [
    ex('1a', 'Sprints', '4×20yd, 2×30yd', ['95%']*5 + ['2×20yd']),
    ex('2a', 'Approach Jumps', '5×2', ['max intent']*5 + ['3×2 deload']),
    ex('3a', 'Depth Jump (mid/high box)', '3×5', ['24 in', '24 in', '30 in', '30 in', '30 in', '18 in 2×5']),
    ex('4a', 'Hang Power Clean', '4×3', [pct_wt(72, CLEAN_1RM), pct_wt(75, CLEAN_1RM), pct_wt(75, CLEAN_1RM),
                                           pct_wt(78, CLEAN_1RM), pct_wt(78, CLEAN_1RM), pct_wt(65, CLEAN_1RM)]),
    ex('5a', 'Dunk Attempts (Wk 9+)', '5-10 attempts', ['—', '—', 'BEGIN: 5 max', '8 max', '10 max', '5 max']),
]
b2_wed = [
    ex('1a', 'Weighted Pull-Up (volume)', '3×8', ['+15 lb', '+20', '+20', '+25', '+30', '+15']),
    ex('2a', 'Face Pull', '3×15', ['Light']*6),
    ex('3a', 'Weighted Side Split Hold', '3×60s', ['20 lb', '25', '25', '30', '30', '20'], flex=True),
    ex('3b', 'PNF Contract-Relax Splits', '3 rounds', ['10s on/30s relax']*6, flex=True, superset_b=True),
    ex('4a', 'Pancake Good Morning', '3×8', ['Mod']*5 + ['Light'], flex=True),
    ex('4b', 'Cossack Squat (loaded)', '3×8/side', ['+15 lb']*5 + ['BW'], flex=True, superset_b=True),
]
b2_thu = [
    ex('1a', 'Power Clean', '5×3', [pct_wt(75, CLEAN_1RM), pct_wt(78, CLEAN_1RM), pct_wt(80, CLEAN_1RM),
                                      pct_wt(80, CLEAN_1RM), pct_wt(82, CLEAN_1RM), pct_wt(70, CLEAN_1RM)]),
    ex('2a', 'Speed Squat', '5×2', [pct_wt(65, SQUAT_1RM), pct_wt(68, SQUAT_1RM), pct_wt(70, SQUAT_1RM),
                                      pct_wt(70, SQUAT_1RM), pct_wt(72, SQUAT_1RM), pct_wt(60, SQUAT_1RM)]),
    ex('3a', 'Jump Squat', '4×4', [pct_wt(30, SQUAT_1RM), pct_wt(30, SQUAT_1RM), pct_wt(33, SQUAT_1RM),
                                     pct_wt(33, SQUAT_1RM), pct_wt(35, SQUAT_1RM), pct_wt(25, SQUAT_1RM)]),
    ex('4a', 'RDL', '4×5,5,4,3', ['235 lb', '245', '255', '265', '275', '215']),
    ex('5a', 'Pendulum Squat', '3×8-10', ['Heavy']*5 + ['Mod']),
]
b2_fri = [
    ex('1a', 'Weighted Pull-Up', '5×3', ['+30 lb', '+35', '+40', '+45', '+50', '+25']),
    ex('2a', 'Weighted Dips', '4×5', ['+25 lb', '+30', '+30', '+35', '+35', '+20']),
    ex('3a', 'Chest-Supported Row', '4×6-8', ['Heavy']*5 + ['Mod']),
    ex('3b', 'Landmine Press', '3×10', ['Mod-Hvy']*5 + ['Light'], superset_b=True),
    ex('4a', 'Incline DB Curl', '3×8-10', ['Mod-Hvy']*5 + ['Light']),
    ex('4b', 'Tricep Pushdown', '3×8-10', ['Mod-Hvy']*5 + ['Light'], superset_b=True),
]
b2_sat = [
    ex('1a', 'Hang Power Snatch', '5×3', ['60% (115)', '62% (120)', '65% (125)', '67% (130)', '70% (135)', '55% (105)']),
    ex('2a', 'DB Push Press (neutral)', '4×4', ['50 lb', '55', '55', '60', '60', '45']),
    ex('3a', 'Single-Leg Bounds', '3×6/leg', ['Med dist']*5 + ['Easy']),
    ex('4a', 'Sprints (finisher)', '4×20yd, 1×40yd', ['90% effort']*5 + ['2×20yd']),
    ex('5a', 'Copenhagen Lift', '3×20s/side', ['BW']*6),
]

build_block('Block 2', 2, B2_FILL, b2_purpose, b2_goals, b2_split, b2_progression, [
    ('SUNDAY', 'Lower Strength (LIGHT maintenance) + Posterior Chain', b2_sun),
    ('MONDAY', 'Upper Horizontal — Bench Density + Row', b2_mon),
    ('TUESDAY', 'Speed-Strength B — Olympic Variety + Sprint Finisher', b2_sat),
    ('WEDNESDAY', 'Recovery + Deep Flex + Pull-Up Volume', b2_wed),
    ('THURSDAY', 'Speed-Strength A — Olympic Power + Loaded Velocity', b2_thu),
    ('FRIDAY', 'Upper Vertical — Pull-Up Heavy + Dips', b2_fri),
    ('SATURDAY', 'JUMP LAB + Dunk Attempts (Wk 9+) — Depth Jumps', b2_tue),
], display_title='REACTIVE + DUNK WINDOW')

# ============== BLOCK 3 ==============
b3_purpose = (
    'Peak + realize. Volume strips out so accumulated fatigue dissipates and the nervous system can '
    'finally express what Blocks 1-2 built (VJ \"delayed transformation\"). Dunk attempts every Tue. '
    'Test bench 260, pull-up +80×8, dunk women\'s BB. Calories move to maintenance — no peak with a deficit.'
)
b3_goals = [
    ('G1', 'DUNK a women\'s basketball on regulation 10\' rim (filmed)', 'Serves: Arc G1'),
    ('G2', 'Bench 1RM 260 lb AND weighted pull-up +80 lb × 8 reps', 'Serves: Arc G2'),
    ('G3', 'Side split distance reduced ≥ 8 cm cumulative from baseline', 'Serves: Arc G3'),
]
b3_split = [
    ('Sun', 'Active Recovery + Posterior Maintenance (very light)'),
    ('Mon', 'Upper Horizontal — Bench Peak (heavy doubles → singles → 1RM)'),
    ('Tue', 'Speed-Strength B — Olympic Variety + Sprint PRs'),
    ('Wed', 'Recovery + Deep Flex + Pull-Up Volume'),
    ('Thu', 'Speed-Strength A — Olympic Taper + Speed Work'),
    ('Fri', 'Upper Vertical — Pull-Up Peak + Dips'),
    ('Sat', 'JUMP DAY — Max Jump + Dunk Attempts every session (TEST DAY)'),
]
b3_progression = [
    'WEEK SHAPE: Sat = TEST DAY (dunk attempts, max jump). Sun = active recovery only — protect Sat performance.',
    'Sun: VERY LIGHT — no squat. Posterior chain hold + pigeon. Body recovers from Sat jump day, prepares for Mon bench.',
    'Bench Mon: Wks 13-14: 3×3 at 200-205. Wks 15-16: heavy doubles/singles 215-220. Wk 17: taper (singles at 90%). Wk 18: TEST 1RM 260.',
    'Pull-Up Fri: heavy triples +55→+65→+70. Tue 3×3 at +50→+60. Wed 3×8 volume at +30. Wk 18: TEST max reps @ +80 lb (target 8).',
    'Squat: NO straight heavy squats. Light Sun is removed entirely. PAP-style heavy squat REMOVED (no day to do it that doesn\'t conflict with Sat).',
    'Power Clean Thu: 4×2 wave 78→82% (155→165 lb). Drop volume; preserve top intent.',
    'Hang Snatch Tue: 4×2 wave 65→72% (125→140 lb). Variety stays in but volume cut.',
    'Speed Squat Thu: 4×2 wave 70→75%. Jump Squat 3×3 @ 30%.',
    'Depth Jump Sat: Wk 13-14 mid box 3×3. Wk 15-16 high box 2×3. Wk 17 taper (low box 2×5). Wk 18: NO depth jumps before test, just approach.',
    'Dunk attempts Sat: 10+ max-intent attempts every Saturday. Film all. Adjust approach mechanics weekly per VJ §8.',
    'Sprints: Sat 3×20yd (jump day, primary); Tue 3×20yd + 1×40yd. Chase sprint PRs Wks 15-16 then taper.',
    'Upper accessories drop to maintenance: 2-3 sets instead of 3-4. Recovery budget goes to the tests.',
    'Wk 13-14 RPE 7-8. Wk 15-16 RPE 8-9 (peak). Wk 17 TAPER (50% volume, 90% intensity). Wk 18 TEST WEEK.',
    'NUTRITION: maintenance calories from Wk 13. No cut during peak (VJ §9 — aggressive deficits compromise force).',
]

b3_sun = [
    ex('1a', 'Stiff-Legged Deadlift (light)', '3×5', ['225 lb', '225', '225', '215', '195', '175']),
    ex('2a', 'Hand-Supported Bulgarian SS', '2×8/leg', ['40 lb', '40', '40', '35', '30', '25']),
    ex('3a', 'Glute-Ham Raise', '2×8', ['BW', 'BW', 'BW', 'BW', 'BW', 'BW']),
    ex('4a', 'Pigeon Stretch', '3×60s/side', ['—'] * 6, flex=True),
    ex('5a', 'Cossack Squat', '3×8/side', ['BW'] * 6, flex=True),
]
b3_mon = [
    ex('1a', 'Bench Press', 'wave to test',
       ['3×3 @ 200', '3×3 @ 205', '3×2 @ 215', '3×2 @ 220', '3×1 @ 235 (90%)', 'TEST 1RM (260)']),
    ex('2a', 'Barbell Row', '4×5', ['205 lb', '210', '210', '215', '195', '175']),
    ex('3a', 'DB Incline Bench (neutral)', '3×5', ['70 lb', '75', '75', '75', '65', '50']),
    ex('3b', 'Chest-Supported Row', '3×6', ['Heavy']*5 + ['Mod'], superset_b=True),
    ex('4a', 'Cable Fly', '2×10', ['Mod']*6),
    ex('4b', 'DB Lateral Raise', '2×10', ['25 lb']*5 + ['15'], superset_b=True),
]
b3_tue = [
    ex('1a', 'Sprints', '3×20yd', ['100% PR']*4 + ['90%', '2×20yd']),
    ex('2a', 'Approach Jumps', '5-8×1', ['max intent']*5 + ['3×1']),
    ex('3a', 'Depth Jump', '2×3 / taper', ['24 in 3×3', '24 in 3×3', '30 in 2×3', '30 in 2×3', '18 in 2×5', 'omit']),
    ex('4a', 'DUNK ATTEMPTS', '10+ attempts', ['MAX intent, film all']*5 + ['TEST DAY: dunk women\'s BB']),
    ex('5a', 'Hang Power Clean (light)', '3×2', [pct_wt(70, CLEAN_1RM), pct_wt(72, CLEAN_1RM), pct_wt(72, CLEAN_1RM),
                                                   pct_wt(70, CLEAN_1RM), pct_wt(65, CLEAN_1RM), 'omit']),
]
b3_wed = [
    ex('1a', 'Weighted Pull-Up (volume)', '3×8', ['+30 lb', '+30', '+30', '+25', '+20', '+15']),
    ex('2a', 'Face Pull', '3×15', ['Light']*6),
    ex('3a', 'Weighted Side Split Hold', '3×90s', ['30 lb', '35', '35', '35', '30', '25'], flex=True),
    ex('3b', 'PNF Contract-Relax Splits', '4 rounds', ['10s on/30s relax']*6, flex=True, superset_b=True),
    ex('4a', 'Pancake Good Morning', '3×8', ['Mod']*6, flex=True),
    ex('4b', 'Cossack Squat (loaded)', '3×6/side', ['+20 lb']*5 + ['BW'], flex=True, superset_b=True),
]
b3_thu = [
    ex('1a', 'Power Clean', '4×2', [pct_wt(78, CLEAN_1RM), pct_wt(80, CLEAN_1RM), pct_wt(82, CLEAN_1RM),
                                      pct_wt(82, CLEAN_1RM), pct_wt(80, CLEAN_1RM), pct_wt(70, CLEAN_1RM)]),
    ex('2a', 'Speed Squat', '4×2', [pct_wt(70, SQUAT_1RM), pct_wt(72, SQUAT_1RM), pct_wt(75, SQUAT_1RM),
                                      pct_wt(75, SQUAT_1RM), pct_wt(70, SQUAT_1RM), pct_wt(60, SQUAT_1RM)]),
    ex('3a', 'Jump Squat', '3×3', [pct_wt(30, SQUAT_1RM)]*5 + [pct_wt(25, SQUAT_1RM)]),
    ex('4a', 'RDL', '3×5', ['265 lb', '275', '275', '265', '245', '215']),
    ex('5a', 'Pendulum Squat', '2×8', ['Mod-Hvy']*5 + ['Mod']),
]
b3_fri = [
    ex('1a', 'Weighted Pull-Up', 'wave to test',
       ['5×3 @ +55', '5×3 @ +60', '4×3 @ +65', '4×3 @ +70', '3×2 @ +75', 'TEST +80×8']),
    ex('2a', 'Weighted Dips', '3×5', ['+35 lb', '+40', '+40', '+40', '+35', '+25']),
    ex('3a', 'Chest-Supported Row', '3×6', ['Heavy']*5 + ['Mod']),
    ex('3b', 'Landmine Press', '2×10', ['Mod-Hvy']*5 + ['Light'], superset_b=True),
    ex('4a', 'Incline DB Curl', '2×10', ['Mod']*6),
    ex('4b', 'Tricep Pushdown', '2×10', ['Mod']*6, superset_b=True),
]
b3_sat = [
    ex('1a', 'Hang Power Snatch', '4×2', ['65% (125)', '68% (130)', '70% (135)', '72% (140)', '68% (130)', '60% (115)']),
    ex('2a', 'DB Push Press (neutral)', '4×3', ['55 lb', '60', '60', '60', '55', '45']),
    ex('3a', 'Continuous Broad Jumps', '3×3', ['Max dist']*5 + ['Med']),
    ex('4a', 'Sprints (PR chase)', '3×20yd, 1×40yd', ['100% PR']*4 + ['90%', '2×20yd']),
    ex('5a', 'Copenhagen Lift', '3×20s/side', ['BW']*6),
]

build_block('Block 3', 3, B3_FILL, b3_purpose, b3_goals, b3_split, b3_progression, [
    ('SUNDAY', 'Active Recovery + Posterior Maintenance (very light)', b3_sun),
    ('MONDAY', 'Upper Horizontal — Bench Peak', b3_mon),
    ('TUESDAY', 'Speed-Strength B — Olympic Variety + Sprint PRs', b3_sat),
    ('WEDNESDAY', 'Recovery + Deep Flex + Pull-Up Volume', b3_wed),
    ('THURSDAY', 'Speed-Strength A — Olympic Taper + Speed Work', b3_thu),
    ('FRIDAY', 'Upper Vertical — Pull-Up Peak + Dips', b3_fri),
    ('SATURDAY', 'JUMP DAY — Max Jump + Dunk Attempts (TEST DAY)', b3_tue),
], display_title='PEAK + REALIZE + TEST')

# ============== Save ==============
# ============== Markdown generation ==============
import os
from datetime import date, timedelta

ARC_START = date(2026, 5, 3)  # Sunday
BUNDLE_DIR = '/Users/andylee/Projects/train/docs/athletes/andy/arc-2026-summer-dunk'
BLOCKS_DIR = f'{BUNDLE_DIR}/blocks'
WEEKS_DIR = f'{BUNDLE_DIR}/weeks'
ACTIVE_DIR = f'{BUNDLE_DIR}/active'
OUTPUTS_DIR = f'{BUNDLE_DIR}/outputs'

for d in (BUNDLE_DIR, BLOCKS_DIR, WEEKS_DIR, ACTIVE_DIR, OUTPUTS_DIR):
    os.makedirs(d, exist_ok=True)

output_path = f'{OUTPUTS_DIR}/hybrid-athletic-plan-v6.xlsx'
wb.save(output_path)
print(f'Wrote {output_path}')

WEEK_META = {
    1:  ('build', 'Establish baselines + groove patterns. Conservative loads.'),
    2:  ('build', 'Small intensity bump. Push bar speed on speed squat + jump squat.'),
    3:  ('build', 'RPE creeps to 7-8. First feel of real load.'),
    4:  ('build', 'Push primary loads. Reactive volume holds.'),
    5:  ('intensify', 'Peak Block 1. RPE 8-9. Test Sat jump quality.'),
    6:  ('deload', 'Deload + 3RM bench test. Re-baseline.'),
    7:  ('build', 'Reactive emphasis begins. Depth jumps introduced (mid-box, 24 in).'),
    8:  ('build', 'Bench density Wk 1 of 3. Depth jumps continue.'),
    9:  ('intensify', 'DUNK ATTEMPTS BEGIN (Sat). High-box depth jumps (30 in).'),
    10: ('intensify', 'Max-intent jumps + density work converging.'),
    11: ('intensify', 'Peak Block 2. RPE 8-9. Last chance to push loads.'),
    12: ('test', 'Deload + Bench 1RM test (target 255).'),
    13: ('build', 'Peak phase begins. Volume strips, intensity holds.'),
    14: ('build', 'Heavy doubles on bench. Pull-up triples climb.'),
    15: ('intensify', 'Peak intensities. Sprint PRs. Dunk attempts every Sat.'),
    16: ('intensify', 'Final peak. Watch for fatigue accumulation.'),
    17: ('deload', 'Taper week — 50% volume, 90% intensity. Protect Sat test.'),
    18: ('test', 'TEST WEEK — DUNK + Bench 260 + Pull-up +80×8.'),
}

BLOCK_SPECS = [
    {
        'block_num': 1,
        'display_title': 'POWER CONVERSION + UPPER BUILD',
        'filename': '2026-05-block-01-power-conversion.md',
        'purpose': b1_purpose,
        'goals': b1_goals,
        'split': b1_split,
        'progression': b1_progression,
        'days': [
            ('SUNDAY', 'Lower Strength (LIGHT maintenance) + Posterior', b1_sun),
            ('MONDAY', 'Upper Horizontal — Bench + Row', b1_mon),
            ('TUESDAY', 'Speed-Strength B — Olympic Variety + Sprint Finisher', b1_sat),
            ('WEDNESDAY', 'Recovery + Deep Flex (Optional) + Pull-Up Volume', b1_wed),
            ('THURSDAY', 'Speed-Strength A — Olympic Power + Speed Squat', b1_thu),
            ('FRIDAY', 'Upper Vertical — Pull-Up + Dips', b1_fri),
            ('SATURDAY', 'JUMP LAB — Sprint + Approach + Depth + Loaded Power', b1_tue),
        ],
    },
    {
        'block_num': 2,
        'display_title': 'REACTIVE + DUNK WINDOW',
        'filename': '2026-06-block-02-reactive-dunk-window.md',
        'purpose': b2_purpose,
        'goals': b2_goals,
        'split': b2_split,
        'progression': b2_progression,
        'days': [
            ('SUNDAY', 'Lower Strength (LIGHT maintenance) + Posterior Chain', b2_sun),
            ('MONDAY', 'Upper Horizontal — Bench Density + Row', b2_mon),
            ('TUESDAY', 'Speed-Strength B — Olympic Variety + Sprint Finisher', b2_sat),
            ('WEDNESDAY', 'Recovery + Deep Flex + Pull-Up Volume', b2_wed),
            ('THURSDAY', 'Speed-Strength A — Olympic Power + Loaded Velocity', b2_thu),
            ('FRIDAY', 'Upper Vertical — Pull-Up Heavy + Dips', b2_fri),
            ('SATURDAY', 'JUMP LAB + Dunk Attempts (Wk 9+) — Depth Jumps', b2_tue),
        ],
    },
    {
        'block_num': 3,
        'display_title': 'PEAK + REALIZE + TEST',
        'filename': '2026-07-block-03-peak-realize-test.md',
        'purpose': b3_purpose,
        'goals': b3_goals,
        'split': b3_split,
        'progression': b3_progression,
        'days': [
            ('SUNDAY', 'Active Recovery + Posterior Maintenance (very light)', b3_sun),
            ('MONDAY', 'Upper Horizontal — Bench Peak', b3_mon),
            ('TUESDAY', 'Speed-Strength B — Olympic Variety + Sprint PRs', b3_sat),
            ('WEDNESDAY', 'Recovery + Deep Flex + Pull-Up Volume', b3_wed),
            ('THURSDAY', 'Speed-Strength A — Olympic Taper + Speed Work', b3_thu),
            ('FRIDAY', 'Upper Vertical — Pull-Up Peak + Dips', b3_fri),
            ('SATURDAY', 'JUMP DAY — Max Jump + Dunk Attempts (TEST DAY)', b3_tue),
        ],
    },
]


def render_block_md(spec):
    L = []
    L.append(f'# Block {spec["block_num"]:02d} — {spec["display_title"]}')
    L.append('')
    L.append('## Purpose')
    L.append('')
    L.append(spec['purpose'])
    L.append('')
    L.append('## Goals')
    L.append('')
    for i, (gid, gtext, serves) in enumerate(spec['goals'], 1):
        L.append(f'{i}. **{gid}**: {gtext} ({serves})')
    L.append('')
    L.append('## Programming Strategy')
    L.append('')
    for i, line in enumerate(spec['progression'], 1):
        L.append(f'{i}. {line}')
    L.append('')
    L.append('## Split (6 Workouts / Week, Wed Optional)')
    L.append('')
    for i, (day_abbr, workout_name) in enumerate(spec['split'], 1):
        L.append(f'{i}. {day_abbr}: {workout_name}')
    L.append('')
    L.append('## Week-by-Week Programming')
    L.append('')
    for week_idx in range(6):
        label = f'Week {week_idx + 1} (Deload / Test)' if week_idx == 5 else f'Week {week_idx + 1}'
        L.append(f'### {label}')
        L.append('')
        for day_name, day_title, exercises in spec['days']:
            L.append(f'#### {day_name} — {day_title}')
            L.append('')
            for i, ex in enumerate(exercises, 1):
                load = ex['weeks'][week_idx]
                L.append(f'{i}. {ex["exercise"]}: {ex["setsreps"]} @ {load}')
            L.append('')
    return '\n'.join(L)


def render_week_md(arc_week, spec, week_idx):
    week_start = ARC_START + timedelta(weeks=arc_week - 1)
    week_end = week_start + timedelta(days=6)
    week_id = f'2026-Arc-W{arc_week:02d}'
    week_type, primary_focus = WEEK_META[arc_week]
    L = []
    L.append(f'# Week {week_id}')
    L.append('')
    L.append('## Week Header')
    L.append('')
    L.append(f'- Block: athletes/andy/blocks/{spec["filename"]}')
    L.append(f'- Block Display: {spec["display_title"]}')
    L.append(f'- Arc Week: {arc_week} of 18')
    L.append(f'- Block Week: {week_idx + 1} of 6')
    L.append(f'- Date Range: {week_start.strftime("%a %b %d")} – {week_end.strftime("%a %b %d, %Y")}')
    L.append(f'- Week Type: {week_type}')
    L.append(f'- Primary Focus: {primary_focus}')
    L.append('')
    L.append('## Block Goals (Why This Week Exists)')
    L.append('')
    for i, (gid, gtext, serves) in enumerate(spec['goals'], 1):
        L.append(f'{i}. **{gid}**: {gtext} ({serves})')
    L.append('')
    L.append('## 7-Day Overview')
    L.append('')
    day_dates = [(week_start + timedelta(days=i)).strftime('%a %b %d') for i in range(7)]
    for i, ((day_name, day_title, _), date_str) in enumerate(zip(spec['days'], day_dates), 1):
        L.append(f'{i}. {date_str} ({day_name}): {day_title}')
    L.append('')

    for (day_name, day_title, exercises), date_str in zip(spec['days'], day_dates):
        L.append(f'## {date_str} ({day_name}) — {day_title}')
        L.append('')
        for i, ex in enumerate(exercises, 1):
            load = ex['weeks'][week_idx]
            L.append(f'{i}. {ex["exercise"]}: {ex["setsreps"]} @ {load}')
        L.append('')

    L.append('## Daily Flexibility (15-20 min post-training)')
    L.append('')
    L.append('1. After each session: pigeon stretch 60s/side, frog 60s, couch stretch 60s/side, 90-90 switches.')
    L.append('2. Wed deep work: weighted side-split holds, PNF contract-relax, pancake good morning.')
    L.append('3. **Never load splits before Sat jump day** (loaded stretch saps adductor force 24-48h).')
    L.append('')
    L.append('## Recovery + Nutrition Targets')
    L.append('')
    L.append('1. Sleep: 7.5-9.0 hours/night.')
    L.append('2. Protein: ≥190 g/day.')
    L.append('3. Hydration: 3.0-4.0 L/day.')
    if spec['block_num'] == 3:
        L.append('4. Bodyweight: MAINTENANCE — no cut during peak (VJ §9: deficits compromise force).')
    else:
        L.append('4. Bodyweight: ~300 cal deficit, target -1 lb/wk. Pause cut if squat top set drops >5%.')
    L.append('')
    L.append('## Logging Checklist')
    L.append('')
    L.append('1. All working sets logged with load, reps, RPE, and notes.')
    L.append('2. Top set logged for primary lifts (squat, bench, pull-up, power clean).')
    L.append('3. Sat jump day: log sprint times, best jump height, dunk attempt count + makes.')
    L.append('4. Daily bodyweight + Sun side-split measurement (cm hip-to-floor).')
    L.append('5. Shoulder + wrist pain rating (1-10) after Mon and Fri sessions.')
    L.append('')
    L.append('## End-of-Week Review (filled Sun)')
    L.append('')
    L.append('1. Wins: <what progressed>')
    L.append('2. Misses: <what was not completed or not clean>')
    L.append('3. Metric Changes: <actual vs target>')
    L.append('4. Adjustment For Next Week: <load/volume/recovery changes>')
    L.append('')
    return '\n'.join(L)


# Generate block files
for spec in BLOCK_SPECS:
    md = render_block_md(spec)
    path = f'{BLOCKS_DIR}/{spec["filename"]}'
    with open(path, 'w') as f:
        f.write(md)
    print(f'Wrote {path}')

# Generate week files (all 18)
for arc_week in range(1, 19):
    block_num = (arc_week - 1) // 6 + 1
    week_idx = (arc_week - 1) % 6
    spec = BLOCK_SPECS[block_num - 1]
    md = render_week_md(arc_week, spec, week_idx)
    week_id = f'2026-Arc-W{arc_week:02d}'
    path = f'{WEEKS_DIR}/{week_id}.md'
    with open(path, 'w') as f:
        f.write(md)
    print(f'Wrote {path}')

# Active snapshots: copy of Block 1 + Week 1
b1_md = render_block_md(BLOCK_SPECS[0])
w1_md = render_week_md(1, BLOCK_SPECS[0], 0)
with open(f'{ACTIVE_DIR}/current-block.md', 'w') as f:
    f.write(b1_md)
print(f'Wrote {ACTIVE_DIR}/current-block.md')
with open(f'{ACTIVE_DIR}/current-week.md', 'w') as f:
    f.write(w1_md)
print(f'Wrote {ACTIVE_DIR}/current-week.md')

